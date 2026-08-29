"""
Generates settlement_currency_review.xlsx: a closed-form, formulas-only Excel
replica of the CIP forward projection + PresentValueFX chain, for human review
against review/settlement_currency_review.py (the marimo notebook) and against
openspec/specs/settlement-currency-present-value/spec.md.

Run with: uv run python build_excel_review.py
"""
import qcfinancial as qcf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

NODE_TS = [1, 30, 90, 180, 365, 730]
NOTIONAL_RATES = [0.01, 0.015, 0.02, 0.025, 0.03, 0.035]
SETTLEMENT_RATES = [0.02, 0.025, 0.03, 0.035, 0.04, 0.045]  # also used as discount curve
SPOT = 900.0
NOMINAL = 1_000_000.0
T_FLOATING = 366
T_FIXED = 183
HISTORICAL_FIX = 875.0

TYPE_SHEETS = [
    ("FixedRate", "FixedRateMultiCurrencyCashflow"),
    ("Ibor", "IborMultiCurrencyCashflow"),
    ("Overnight", "OvernightIndexMultiCurrencyCashflow"),
    ("CompOvernight", "CompoundedOvernightRateMultiCurrencyCashflow2"),
    ("Simple_NDF", "SimpleMultiCurrencyCashflow"),
]

HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
SECTION_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BOLD = Font(bold=True)

# Rows 2..RESULTS_LAST_ROW hold the live qcfinancial results table at the top of each type
# sheet; the formula-derivation section below starts at ROW_OFFSET so it never overlaps.
RESULTS_LAST_ROW = 11
ROW_OFFSET = 13


def col(i):
    """0-based node index (0..5) -> Excel column letter, offset to start at B (A is for labels)."""
    return get_column_letter(2 + i)


def _build_market():
    usd = qcf.QCUSD()
    clp = qcf.QCCLP()
    fx_rate = qcf.FXRate(usd, clp)  # USD strong, CLP weak
    one_day = qcf.Tenor("1D")
    fx_calendar = qcf.BusinessCalendar(qcf.QCDate(1, 1, 2021), 20)
    fx_index = qcf.FXRateIndex(fx_rate, "USDCLPOBS", one_day, one_day, fx_calendar)
    lin_act360 = qcf.QCInterestRate(0.0, qcf.QCAct360(), qcf.QCLinearWf())
    sofr_index = qcf.InterestRateIndex(
        "SOFRRATE", lin_act360, one_day, one_day, fx_calendar, fx_calendar, usd)
    return usd, clp, fx_index, lin_act360, sofr_index


def _build_curve(lin_act360, rates):
    plazos = qcf.long_vec()
    for p in NODE_TS:
        plazos.append(p)
    rate_values = qcf.double_vec()
    for r in rates:
        rate_values.append(r)
    curve = qcf.QCCurve(plazos, rate_values)
    interp = qcf.QCLinearInterpolator(curve)
    return qcf.ZeroCouponCurve(interp, lin_act360)


def _make_cf(sheet_key, usd, clp, fx_index, lin_act360, sofr_index, start, end, fx_fix):
    if sheet_key == "FixedRate":
        return qcf.FixedRateMultiCurrencyCashflow(
            start, end, end, NOMINAL, NOMINAL, True, lin_act360, usd, fx_fix, clp, fx_index, 1.0)
    if sheet_key == "Ibor":
        cf = qcf.IborMultiCurrencyCashflow(
            sofr_index, start, end, start, end, NOMINAL, NOMINAL, True, usd, 0.0, 1.0,
            fx_fix, clp, fx_index, 1.0)
        cf.set_interest_rate_value(0.0)
        return cf
    if sheet_key == "Overnight":
        return qcf.OvernightIndexMultiCurrencyCashflow(
            start, end, start, end, end, usd, NOMINAL, NOMINAL, True, 0.0, 1.0,
            lin_act360, "SOFRINDEX", 8, qcf.DatesForEquivalentRate.ACCRUAL, fx_fix, clp, fx_index)
    if sheet_key == "CompOvernight":
        return qcf.CompoundedOvernightRateMultiCurrencyCashflow2(
            sofr_index, start, end, end, qcf.DateList(), NOMINAL, NOMINAL, True,
            usd, 0.0, 1.0, lin_act360, 8, 0, 0, fx_fix, clp, fx_index)
    if sheet_key == "Simple_NDF":
        return qcf.SimpleMultiCurrencyCashflow(end, NOMINAL, usd, fx_fix, clp, fx_index, 1.0)
    raise ValueError(f"unknown sheet_key {sheet_key!r}")


def compute_actual_results(sheet_key):
    """Runs the real qcfinancial pipeline for one cashflow type — the same three steps as
    settlement_currency_review.py's run_demo(): spot fixing, CIP forward (floating), CIP
    forward (already fixed). Returns plain Python values, written into the sheet as literals
    so a reader sees what qcfinancial actually produced, not just the Excel replica of it."""
    usd, clp, fx_index, lin_act360, sofr_index = _build_market()
    notional_curve = _build_curve(lin_act360, NOTIONAL_RATES)
    settlement_curve = _build_curve(lin_act360, SETTLEMENT_RATES)

    start = qcf.QCDate(2, 1, 2024)
    end = qcf.QCDate(2, 1, 2025)
    fx_fix = qcf.QCDate(2, 7, 2024)
    val_floating = qcf.QCDate(2, 1, 2024)
    val_fixed = qcf.QCDate(3, 7, 2024)

    # 1. Spot fixing.
    cf1 = _make_cf(sheet_key, usd, clp, fx_index, lin_act360, sofr_index, start, end, fx_fix)
    ts = qcf.time_series()
    ts[fx_fix] = HISTORICAL_FIX
    estimator = qcf.FXRateEstimator(ts, -1.0)
    fixed_cf = qcf.ForwardFXRates().set_fx_rate(val_fixed, cf1, estimator)
    settlement_amount_fixing = fixed_cf.settlement_amount()

    # 2. CIP forward, still floating.
    cf2 = _make_cf(sheet_key, usd, clp, fx_index, lin_act360, sofr_index, start, end, fx_fix)
    projected_floating = qcf.ForwardFXRates().set_fx_rate_cip(
        val_floating, SPOT, cf2, notional_curve, settlement_curve)
    pvfx_floating = qcf.PresentValueFX()
    pv_floating = pvfx_floating.pv(val_floating, projected_floating, settlement_curve)

    # 3. CIP forward, already fixed.
    cf3 = _make_cf(sheet_key, usd, clp, fx_index, lin_act360, sofr_index, start, end, fx_fix)
    cf3.set_fx_rate_index_value(HISTORICAL_FIX)
    projected_fixed = qcf.ForwardFXRates().set_fx_rate_cip(
        val_fixed, SPOT, cf3, notional_curve, settlement_curve)
    pv_fixed = qcf.PresentValueFX().pv(val_fixed, projected_fixed, settlement_curve)

    return {
        "settlement_amount_fixing": settlement_amount_fixing,
        "pv_floating": pv_floating,
        "fx_delta_floating": pvfx_floating.get_fx_delta(),
        "notional_deriv_node5": pvfx_floating.get_notional_curve_derivatives()[4],
        "cip_deriv_node5": pvfx_floating.get_cip_settlement_curve_derivatives()[4],
        "disc_deriv_node5": pvfx_floating.get_discount_curve_derivatives()[4],
        "pv_fixed": pv_fixed,
    }


def build_curves_sheet(wb):
    ws = wb.active
    ws.title = "Curves"
    ws["A1"] = "Node index"
    ws["A2"] = "Node t (days)"
    ws["A3"] = "Notional curve rate"
    ws["A4"] = "Settlement/discount curve rate"
    for r in (1, 2, 3, 4):
        ws.cell(row=r, column=1).font = BOLD

    for i in range(6):
        c = col(i)
        ws[f"{c}1"] = i + 1
        ws[f"{c}2"] = NODE_TS[i]
        ws[f"{c}3"] = NOTIONAL_RATES[i]
        ws[f"{c}4"] = SETTLEMENT_RATES[i]

    labels = [
        ("A6", "Spot FX (USDCLP)", "B6", SPOT),
        ("A7", "Nominal (USD)", "B7", NOMINAL),
        ("A8", "t_floating (days to end date)", "B8", T_FLOATING),
        ("A9", "t_fixed (days to end date)", "B9", T_FIXED),
        ("A10", "Historical FX fixing", "B10", HISTORICAL_FIX),
    ]
    for label_cell, label, value_cell, value in labels:
        ws[label_cell] = label
        ws[label_cell].font = BOLD
        ws[value_cell] = value

    ws["A12"] = (
        "Interpolation: linear on the rate between bracketing nodes "
        "(QCLinearInterpolator). wf(t) = 1 + rate(t) * t/360 (QCAct360 + QCLinearWf). "
        "DF(t) = 1/wf(t). Matches ZeroCouponCurve::getDiscountFactorAt exactly."
    )
    ws.column_dimensions["A"].width = 34
    for i in range(6):
        ws.column_dimensions[col(i)].width = 12
    return ws


def _write_results_block(ws, sheet_key, actual):
    """Rows 2..RESULTS_LAST_ROW: the actual qcfinancial output, as plain values (not
    formulas) — read this first to see what the answer is; the formula section below
    (starting at ROW_OFFSET) shows how it's derived, cell by cell."""
    ws["A2"] = f"qcfinancial results (Python, live — see settlement_currency_review.py)"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].fill = RESULT_FILL
    for c in "BCDEF":
        ws[f"{c}2"].fill = RESULT_FILL

    headers = ["Step", "Amount / PV (CLP)", "Notes"]
    for c, h in zip("BCD", headers):
        ws[f"{c}3"] = h
        ws[f"{c}3"].font = BOLD

    rows = [
        ("Spot fixing (historical FX = 875.0)",
         actual["settlement_amount_fixing"],
         "settlement_amount() after ForwardFXRates.set_fx_rate"),
        ("CIP forward, still floating",
         actual["pv_floating"],
         f"PresentValueFX.pv; FX delta = {actual['fx_delta_floating']:,.4f}"),
        ("CIP forward, already fixed",
         actual["pv_fixed"],
         "PresentValueFX.pv; all curve/spot derivatives are 0 once fixed"),
    ]
    r = 4
    for label, value, note in rows:
        ws[f"A{r}"] = label
        ws[f"B{r}"] = value
        ws[f"B{r}"].number_format = "#,##0.0000"
        ws[f"D{r}"] = note
        r += 1

    r += 1
    ws[f"A{r}"] = "Floating-case curve-vertex derivatives at node 5 (t=365d):"
    ws[f"A{r}"].font = BOLD
    r += 1
    ws[f"A{r}"] = "  notional-curve"
    ws[f"B{r}"] = actual["notional_deriv_node5"]
    ws[f"B{r}"].number_format = "#,##0.0000"
    r += 1
    ws[f"A{r}"] = "  CIP-settlement-curve"
    ws[f"B{r}"] = actual["cip_deriv_node5"]
    ws[f"B{r}"].number_format = "#,##0.0000"
    r += 1
    ws[f"A{r}"] = "  discount-curve"
    ws[f"B{r}"] = actual["disc_deriv_node5"]
    ws[f"B{r}"].number_format = "#,##0.0000"

    assert r <= RESULTS_LAST_ROW, f"results block overflowed row {r} > {RESULTS_LAST_ROW} for {sheet_key}"


def build_type_sheet(wb, sheet_key, type_name):
    ws = wb.create_sheet(sheet_key)
    ws["A1"] = type_name
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 34
    for i in range(6):
        ws.column_dimensions[col(i)].width = 14
    ws.column_dimensions["D"].width = 46

    _write_results_block(ws, sheet_key, compute_actual_results(sheet_key))

    CV = "Curves"

    def node_range(row):
        return f"{CV}!$B${row}:$G${row}"

    idx_range = node_range(1)
    t_range = node_range(2)
    notional_rate_range = node_range(3)
    settlement_rate_range = node_range(4)

    r = 3 + ROW_OFFSET
    NODE_T_ROW = r
    ws[f"A{r}"] = "Node t (days)"
    for i in range(6):
        ws[f"{col(i)}{r}"] = f"={CV}!{col(i)}2"
    r += 1
    ws[f"A{r}"] = "Notional curve rate"
    for i in range(6):
        ws[f"{col(i)}{r}"] = f"={CV}!{col(i)}3"
    r += 1
    ws[f"A{r}"] = "Settlement/discount curve rate"
    for i in range(6):
        ws[f"{col(i)}{r}"] = f"={CV}!{col(i)}4"
    for row in (3 + ROW_OFFSET, 4 + ROW_OFFSET, 5 + ROW_OFFSET):
        ws[f"A{row}"].font = BOLD

    # --- Scalar market inputs -------------------------------------------------
    r = 7 + ROW_OFFSET
    ws[f"A{r}"] = "Spot"
    ws[f"B{r}"] = f"={CV}!B6"
    r += 1
    ws[f"A{r}"] = "Nominal (notional-currency amount(); zero-rate construction, see notebook)"
    ws[f"B{r}"] = f"={CV}!B7"
    r += 1
    ws[f"A{r}"] = "t_floating"
    ws[f"B{r}"] = f"={CV}!B8"
    r += 1
    ws[f"A{r}"] = "t_fixed"
    ws[f"B{r}"] = f"={CV}!B9"
    r += 1
    ws[f"A{r}"] = "Historical FX fixing"
    ws[f"B{r}"] = f"={CV}!B10"
    (SPOT_R, NOMINAL_R, TFLOAT_R, TFIXED_R, HISTFIX_R) = (
        7 + ROW_OFFSET, 8 + ROW_OFFSET, 9 + ROW_OFFSET, 10 + ROW_OFFSET, 11 + ROW_OFFSET)

    # --- Floating-case CIP forward projection --------------------------------
    r = 13 + ROW_OFFSET
    ws[f"A{r}"] = "CIP forward projection (still floating)"
    ws[f"A{r}"].fill = SECTION_FILL
    ws[f"A{r}"].font = BOLD

    r = 14 + ROW_OFFSET
    ws[f"A{r}"] = "idx (largest node <= t_floating)"
    ws[f"B{r}"] = f"=MATCH(B{TFLOAT_R},{t_range},1)"
    IDX_R = r
    r += 1
    ws[f"A{r}"] = "x1, x2 (bracketing node t's)"
    ws[f"B{r}"] = f"=INDEX({t_range},B{IDX_R})"
    ws[f"C{r}"] = f"=IF(B{IDX_R}={len(NODE_TS)},B{r},INDEX({t_range},B{IDX_R}+1))"
    X1_R, X2_R = r, r
    r += 1
    ws[f"A{r}"] = "frac = (t_floating - x1) / (x2 - x1)"
    ws[f"B{r}"] = f"=IF(B{IDX_R}={len(NODE_TS)},0,(B{TFLOAT_R}-B{X1_R})/(C{X2_R}-B{X1_R}))"
    FRAC_R = r
    r += 1
    ws[f"A{r}"] = "Interpolated notional-curve rate at t_floating"
    ws[f"B{r}"] = f"=INDEX({notional_rate_range},B{IDX_R})+(INDEX({notional_rate_range},B{IDX_R}+1-IF(B{IDX_R}={len(NODE_TS)},1,0))-INDEX({notional_rate_range},B{IDX_R}))*B{FRAC_R}"
    RATE_N_R = r
    r += 1
    ws[f"A{r}"] = "Interpolated settlement-curve rate at t_floating"
    ws[f"B{r}"] = f"=INDEX({settlement_rate_range},B{IDX_R})+(INDEX({settlement_rate_range},B{IDX_R}+1-IF(B{IDX_R}={len(NODE_TS)},1,0))-INDEX({settlement_rate_range},B{IDX_R}))*B{FRAC_R}"
    RATE_S_R = r
    r += 1
    ws[f"A{r}"] = "wf_notional = 1 + rate * t_floating/360"
    ws[f"B{r}"] = f"=1+B{RATE_N_R}*B{TFLOAT_R}/360"
    WF_N_R = r
    r += 1
    ws[f"A{r}"] = "wf_settlement = 1 + rate * t_floating/360"
    ws[f"B{r}"] = f"=1+B{RATE_S_R}*B{TFLOAT_R}/360"
    WF_S_R = r
    r += 1
    ws[f"A{r}"] = "DF_notional(t_floating) = 1/wf_notional"
    ws[f"B{r}"] = f"=1/B{WF_N_R}"
    DFN_R = r
    r += 1
    ws[f"A{r}"] = "DF_settlement(t_floating) = 1/wf_settlement"
    ws[f"B{r}"] = f"=1/B{WF_S_R}"
    DFS_R = r
    r += 1
    ws[f"A{r}"] = "Forward = Spot * DF_notional / DF_settlement"
    ws[f"B{r}"] = f"=B{SPOT_R}*B{DFN_R}/B{DFS_R}"
    FWD_R = r
    r += 1
    ws[f"A{r}"] = "Amount (settlement ccy, strong-side notional) = Nominal * Forward"
    ws[f"B{r}"] = f"=B{NOMINAL_R}*B{FWD_R}"
    AMT_R = r
    r += 1
    ws[f"A{r}"] = "PV (floating) = Amount * DF_settlement  [discount curve = settlement curve here]"
    ws[f"B{r}"] = f"=B{AMT_R}*B{DFS_R}"
    PV_FLOAT_R = r
    r += 1
    ws[f"A{r}"] = "dForward/dSpot = DF_notional/DF_settlement"
    ws[f"B{r}"] = f"=B{DFN_R}/B{DFS_R}"
    DFWD_DSPOT_R = r
    r += 1
    ws[f"A{r}"] = "FX delta = Nominal * dForward/dSpot * DF_settlement"
    ws[f"B{r}"] = f"=B{NOMINAL_R}*B{DFWD_DSPOT_R}*B{DFS_R}"
    FXDELTA_R = r

    # --- Per-node curve-vertex derivatives ------------------------------------
    r += 2
    ws[f"A{r}"] = "Per-node curve-vertex derivatives (floating case)"
    ws[f"A{r}"].fill = SECTION_FILL
    ws[f"A{r}"].font = BOLD
    r += 1
    ws[f"A{r}"] = "dRate/dNode_j (shared: same node grid for both curves)"
    DRATE_R = r
    for i in range(6):
        c = col(i)
        # Same-sheet comparison against this row's own "Node t" reference row and the
        # already-computed bracket endpoints (B{X1_R}, C{X2_R}) — avoids the cross-sheet row-1
        # index-array comparison, which one formula-evaluator (used only for this workbook's
        # automated verification, not by Excel itself) resolved off-by-one against a per-column
        # cross-sheet reference; this same-sheet form verified correctly instead.
        ws[f"{c}{r}"] = (
            f"=IF({c}{NODE_T_ROW}=B{X1_R},1-B{FRAC_R},"
            f"IF({c}{NODE_T_ROW}=C{X2_R},B{FRAC_R},0))"
        )
    r += 1
    ws[f"A{r}"] = "dDF_notional/dNode_j = -(t_floating/360) * dRate/dNode_j / wf_notional^2"
    DDFN_R = r
    for i in range(6):
        c = col(i)
        ws[f"{c}{r}"] = f"=-(B{TFLOAT_R}/360)*{c}{DRATE_R}/(B{WF_N_R}^2)"
    r += 1
    ws[f"A{r}"] = "dDF_settlement/dNode_j = -(t_floating/360) * dRate/dNode_j / wf_settlement^2"
    DDFS_R = r
    for i in range(6):
        c = col(i)
        ws[f"{c}{r}"] = f"=-(B{TFLOAT_R}/360)*{c}{DRATE_R}/(B{WF_S_R}^2)"
    r += 1
    ws[f"A{r}"] = "dPV/dNotionalNode_j = Nominal * Spot * dDF_notional/dNode_j / DF_settlement * DF_settlement"
    DPV_DNOTIONAL_R = r
    for i in range(6):
        c = col(i)
        ws[f"{c}{r}"] = f"=B{NOMINAL_R}*B{SPOT_R}*{c}{DDFN_R}/B{DFS_R}*B{DFS_R}"
    r += 1
    ws[f"A{r}"] = "dPV/dCipSettlementNode_j = Nominal * (-Spot*DF_notional/DF_settlement^2) * dDF_settlement/dNode_j * DF_settlement"
    DPV_DCIP_R = r
    for i in range(6):
        c = col(i)
        ws[f"{c}{r}"] = f"=B{NOMINAL_R}*(-B{SPOT_R}*B{DFN_R}/(B{DFS_R}^2))*{c}{DDFS_R}*B{DFS_R}"
    r += 1
    ws[f"A{r}"] = "dPV/dDiscountNode_j = Amount * dDF_settlement/dNode_j  [discount curve = settlement curve]"
    DPV_DDISC_R = r
    for i in range(6):
        c = col(i)
        ws[f"{c}{r}"] = f"=B{AMT_R}*{c}{DDFS_R}"
    r += 1
    ws[f"A{r}"] = "Cancellation check: dPV/dCip + dPV/dDiscount (≈0 expected, strong-side, same curve)"
    CHECK_R = r
    for i in range(6):
        c = col(i)
        ws[f"{c}{r}"] = f"={c}{DPV_DCIP_R}+{c}{DPV_DDISC_R}"

    # --- Already-fixed case ----------------------------------------------------
    r += 2
    ws[f"A{r}"] = "Already-fixed case (valuation date after FX fixing date)"
    ws[f"A{r}"].fill = SECTION_FILL
    ws[f"A{r}"].font = BOLD
    r += 1
    ws[f"A{r}"] = "idx_fixed (largest node <= t_fixed)"
    ws[f"B{r}"] = f"=MATCH(B{TFIXED_R},{t_range},1)"
    IDXX_R = r
    r += 1
    ws[f"A{r}"] = "x1, x2 (bracketing node t's)"
    ws[f"B{r}"] = f"=INDEX({t_range},B{IDXX_R})"
    ws[f"C{r}"] = f"=INDEX({t_range},B{IDXX_R}+1)"
    X1X_R = r
    r += 1
    ws[f"A{r}"] = "frac_fixed"
    ws[f"B{r}"] = f"=(B{TFIXED_R}-B{X1X_R})/(C{X1X_R}-B{X1X_R})"
    FRACX_R = r
    r += 1
    ws[f"A{r}"] = "Interpolated settlement-curve rate at t_fixed"
    ws[f"B{r}"] = f"=INDEX({settlement_rate_range},B{IDXX_R})+(INDEX({settlement_rate_range},B{IDXX_R}+1)-INDEX({settlement_rate_range},B{IDXX_R}))*B{FRACX_R}"
    RATESX_R = r
    r += 1
    ws[f"A{r}"] = "wf_settlement(t_fixed)"
    ws[f"B{r}"] = f"=1+B{RATESX_R}*B{TFIXED_R}/360"
    WFSX_R = r
    r += 1
    ws[f"A{r}"] = "DF_settlement(t_fixed)"
    ws[f"B{r}"] = f"=1/B{WFSX_R}"
    DFSX_R = r
    r += 1
    ws[f"A{r}"] = "Amount_fixed = Nominal * historical fixing"
    ws[f"B{r}"] = f"=B{NOMINAL_R}*B{HISTFIX_R}"
    AMTX_R = r
    r += 1
    ws[f"A{r}"] = "PV (already fixed) = Amount_fixed * DF_settlement(t_fixed)"
    ws[f"B{r}"] = f"=B{AMTX_R}*B{DFSX_R}"
    PV_FIXED_R = r
    r += 1
    ws[f"A{r}"] = "All curve/spot derivatives (already fixed) = 0 (no FX-forward sensitivity once fixed)"

    # --- Cross-check against the qcfinancial results block at the top of this sheet -------
    r += 2
    ws[f"A{r}"] = "Cross-check: this formula derivation vs. the qcfinancial results above (rows 4-6)"
    ws[f"A{r}"].fill = HEADER_FILL
    ws[f"A{r}"].font = BOLD
    r += 1
    ws[f"A{r}"] = "Match (floating)? [formula PV vs. row 5 result]"
    ws[f"B{r}"] = f"=ABS(B{PV_FLOAT_R}-B5)<0.01"
    r += 1
    ws[f"A{r}"] = "Match (already fixed)? [formula PV vs. row 6 result]"
    ws[f"B{r}"] = f"=ABS(B{PV_FIXED_R}-B6)<0.01"

    return ws


def main():
    wb = Workbook()
    build_curves_sheet(wb)
    for sheet_key, type_name in TYPE_SHEETS:
        build_type_sheet(wb, sheet_key, type_name)
    out_path = "settlement_currency_review.xlsx"
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
